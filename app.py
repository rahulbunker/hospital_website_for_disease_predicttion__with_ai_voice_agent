"""
CityMed Hospital - Disease Prediction Backend
Run:  python app.py
Open: http://localhost:5000
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
import csv, os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='.')

# ── File paths ────────────────────────────────────────────────────────────────
DATA_DIR            = 'data'
ALL_LOGINS_FILE     = os.path.join(DATA_DIR, 'all_logins.csv')
POSITIVE_EXCEL      = os.path.join(DATA_DIR, 'positive_patients.xlsx')
POSITIVE_CSV        = os.path.join(DATA_DIR, 'positive_patients.csv')

ALL_LOGIN_HEADERS = ['Date', 'Time', 'Name', 'Mobile', 'Age', 'Gender', 'City']

POSITIVE_HEADERS = [
    'Date', 'Time', 'Name', 'Mobile', 'Age', 'Gender', 'City',
    'Disease', 'Risk %',
    'Input: Glucose/Systolic/Creatinine/SGPT',
    'Input: BMI/Diastolic/Urea/SGOT',
    'Input: HbA1c/Salt/UrineProtein/Bilirubin',
    'Input: FamilyHistory/Stress/Swelling/Alcohol',
    'Input: ChestPain/Symptoms',
    'Advice'
]

os.makedirs(DATA_DIR, exist_ok=True)

def ensure_csv(filepath, headers):
    if not os.path.exists(filepath):
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            csv.writer(f).writerow(headers)

def ensure_excel():
    """Create Excel file with styled header if it doesn't exist."""
    if os.path.exists(POSITIVE_EXCEL):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Positive Patients'
    
    # Style header
    header_fill = PatternFill("solid", fgColor="B71C1C")  # dark red
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col, h in enumerate(POSITIVE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border
    
    # Column widths
    widths = [10, 8, 18, 12, 5, 8, 10, 14, 7, 20, 20, 20, 20, 20, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A2'
    wb.save(POSITIVE_EXCEL)

ensure_csv(ALL_LOGINS_FILE, ALL_LOGIN_HEADERS)
ensure_excel()
ensure_csv(POSITIVE_CSV, POSITIVE_HEADERS)


# ── Helper ────────────────────────────────────────────────────────────────────
def get_inputs_flat(disease, inputs):
    """Map inputs to our 5 flat columns regardless of disease type."""
    i = inputs or {}
    if disease == 'Diabetes':
        return [
            i.get('Glucose',''), i.get('BMI',''), i.get('HbA1c',''),
            i.get('Family History',''), i.get('Symptoms Count','')
        ]
    elif disease == 'Heart Disease':
        return [
            i.get('BP',''), i.get('Cholesterol',''), '',
            '', i.get('Chest Pain','') + ' | ExBreath:' + str(i.get('Exercise Breathlessness',''))
        ]
    elif disease == 'Blood Pressure':
        return [
            i.get('Systolic',''), i.get('Diastolic',''), i.get('Salt',''),
            i.get('Stress',''), f"Symptoms:{i.get('Symptoms','')}"
        ]
    elif disease == 'Kidney Disease':
        return [
            i.get('Creatinine',''), i.get('Blood Urea',''), i.get('Urine Protein',''),
            i.get('Swelling',''), ''
        ]
    elif disease == 'Liver Disease':
        return [
            i.get('SGPT',''), i.get('SGOT',''), i.get('Bilirubin',''),
            i.get('Alcohol',''), f"Symptoms:{i.get('Symptoms','')}"
        ]
    return ['','','','','']


def append_to_excel(row_data):
    """Append one row to the Excel file with alternating row color."""
    wb = openpyxl.load_workbook(POSITIVE_EXCEL)
    ws = wb.active
    next_row = ws.max_row + 1
    
    # Alternating row fill
    fill_color = "FFEBEE" if next_row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill("solid", fgColor=fill_color)
    thin = Side(style='thin', color="EEEEEE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.fill = row_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = border
    
    ws.row_dimensions[next_row].height = 20
    wb.save(POSITIVE_EXCEL)


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/save_login', methods=['POST'])
def save_login():
    """Auto-called when any user logs in — saves to all_logins.csv."""
    data = request.get_json() or {}
    p = data.get('patient', {})
    now = datetime.now()
    row = [
        now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'),
        p.get('name',''), p.get('mobile',''),
        p.get('age',''), p.get('gender',''), p.get('city','')
    ]
    with open(ALL_LOGINS_FILE, 'a', newline='', encoding='utf-8-sig') as f:
        csv.writer(f).writerow(row)
    return jsonify({'success': True})


@app.route('/save_positive', methods=['POST'])
def save_positive():
    """
    Auto-called silently when a disease result is POSITIVE.
    User kuch nahi karta — background mein save hota hai.
    """
    data = request.get_json() or {}
    p = data.get('patient', {})
    res = data.get('result', {})

    now = datetime.now()
    date_str = now.strftime('%d/%m/%Y')
    time_str = now.strftime('%H:%M:%S')

    disease  = res.get('disease', '')
    risk     = round(res.get('risk', 0))
    advice   = res.get('advice', '')
    inputs   = res.get('inputs', {})

    flat_inputs = get_inputs_flat(disease, inputs)

    row = [
        date_str, time_str,
        p.get('name',''), p.get('mobile',''),
        p.get('age',''), p.get('gender',''), p.get('city',''),
        disease, f"{risk}%",
        *flat_inputs,
        advice
    ]

    # Save to CSV
    with open(POSITIVE_CSV, 'a', newline='', encoding='utf-8-sig') as f:
        csv.writer(f).writerow(row)

    # Save to Excel
    try:
        append_to_excel(row)
    except Exception as e:
        print(f"Excel write error: {e}")

    print(f"[AUTO-SAVED] {p.get('name','')} | {p.get('mobile','')} | {disease} | {risk}% risk")
    return jsonify({'success': True})


@app.route('/download/<file_type>')
def download_file(file_type):
    now_str = datetime.now().strftime('%d%m%Y')
    if file_type == 'logins':
        return send_file(ALL_LOGINS_FILE, as_attachment=True,
                         download_name=f'CityMed_All_Logins_{now_str}.csv',
                         mimetype='text/csv')
    elif file_type == 'positive_csv':
        return send_file(POSITIVE_CSV, as_attachment=True,
                         download_name=f'CityMed_Positive_Patients_{now_str}.csv',
                         mimetype='text/csv')
    elif file_type == 'positive_excel':
        return send_file(POSITIVE_EXCEL, as_attachment=True,
                         download_name=f'CityMed_Positive_Patients_{now_str}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return jsonify({'error': 'Invalid type'}), 400


@app.route('/stats')
def stats():
    """Live counts for hospital dashboard."""
    def count_rows(f):
        if not os.path.exists(f): return 0
        with open(f, 'r', encoding='utf-8-sig') as fp:
            return max(0, sum(1 for _ in fp) - 1)

    login_count = count_rows(ALL_LOGINS_FILE)
    pos_count   = count_rows(POSITIVE_CSV)

    # Disease-wise counts from positive CSV
    disease_counts = {}
    if os.path.exists(POSITIVE_CSV):
        with open(POSITIVE_CSV, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                d = row.get('Disease','')
                disease_counts[d] = disease_counts.get(d, 0) + 1

    return jsonify({
        'total_logins': login_count,
        'positive_patients': pos_count,
        'disease_breakdown': disease_counts
    })


if __name__ == '__main__':
    print("=" * 60)
    print("  🏥 CityMed Hospital — Disease Prediction System")
    print("=" * 60)
    print(f"  🌐 Website:           http://localhost:5000")
    print(f"  📋 All logins CSV:    {ALL_LOGINS_FILE}")
    print(f"  🔴 Positive CSV:      {POSITIVE_CSV}")
    print(f"  📊 Positive Excel:    {POSITIVE_EXCEL}  ← MAIN FILE")
    print()
    print(f"  📥 Download logins:   http://localhost:5000/download/logins")
    print(f"  📥 Download positive Excel: http://localhost:5000/download/positive_excel")
    print(f"  📊 Live stats:        http://localhost:5000/stats")
    print()
    print("  ✅ Positive results automatically save — no user action needed!")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)